#!/usr/bin/env python3
"""
Generate London market data files for the Ground Signal app.

This keeps the Berlin source data intact and emits London-specific JSON/TS
artifacts from the two workbook files in the repo plus a small manual anchor
set for retail, schools, and premium OOH placements.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"

CREATIVE_WORKBOOK = ROOT / "London_Creative_Ecosystem_Directory_open_data.xlsx"
OOH_WORKBOOK = ROOT / "London_OOH_Master_List_expanded_public_proxy.xlsx"

LOCATIONS_OUTPUT = DATA_DIR / "london_locations.json"
OOH_OUTPUT = DATA_DIR / "london_ooh_locations.json"
NEIGHBOURHOODS_OUTPUT = DATA_DIR / "london_neighbourhoods.json"
ZONE_METADATA_OUTPUT = DATA_DIR / "londonZoneMetadata.ts"
TRANSIT_WEIGHTS_OUTPUT = DATA_DIR / "londonTransitWeights.ts"
MARKET_OUTPUT = DATA_DIR / "londonMarket.ts"

DATA_REFRESH_DATE = "2026-03-10"

ZONE_ORDER = [
    "West End",
    "King's Cross",
    "Camden",
    "Shoreditch",
    "South Bank",
    "Brixton",
    "Peckham",
    "Stratford",
]

ZONE_POLYGONS = {
    "West End": [
        (51.5050, -0.1530),
        (51.5170, -0.1530),
        (51.5170, -0.1120),
        (51.5050, -0.1120),
    ],
    "King's Cross": [
        (51.5180, -0.1460),
        (51.5370, -0.1460),
        (51.5370, -0.0910),
        (51.5180, -0.0910),
    ],
    "Camden": [
        (51.5320, -0.1650),
        (51.5500, -0.1650),
        (51.5500, -0.1230),
        (51.5320, -0.1230),
    ],
    "Shoreditch": [
        (51.5150, -0.1020),
        (51.5320, -0.1020),
        (51.5320, -0.0450),
        (51.5150, -0.0450),
    ],
    "South Bank": [
        (51.4950, -0.1340),
        (51.5130, -0.1340),
        (51.5130, -0.0650),
        (51.4950, -0.0650),
    ],
    "Brixton": [
        (51.4510, -0.1530),
        (51.4700, -0.1530),
        (51.4700, -0.1050),
        (51.4510, -0.1050),
    ],
    "Peckham": [
        (51.4550, -0.0910),
        (51.4850, -0.0910),
        (51.4850, -0.0300),
        (51.4550, -0.0300),
    ],
    "Stratford": [
        (51.5320, -0.0350),
        (51.5500, -0.0350),
        (51.5500, 0.0260),
        (51.5320, 0.0260),
    ],
}

ZONE_COPY = {
    "West End": {
        "brandFit": "high",
        "description": "London's flagship visibility zone. Dense gallery coverage, major retail anchors, and premium OOH make it the clearest market-entry stage for high-impact launches.",
    },
    "King's Cross": {
        "brandFit": "high",
        "description": "Design-campus corridor with strong commuter flow. Balances cultural institutions, creative workspaces, and education touchpoints for launch programming with substance.",
    },
    "Camden": {
        "brandFit": "high",
        "description": "Youth and music energy cluster. Best suited to guerrilla work, nightlife adjacency, and creator-facing activations with strong street relevance.",
    },
    "Shoreditch": {
        "brandFit": "high",
        "description": "London's sharpest creator-network zone. Agency density, coworking, and cultural nightlife give it the strongest seeding and community-amplification profile.",
    },
    "South Bank": {
        "brandFit": "medium",
        "description": "Public-culture spine with broad footfall. Museums, theatres, and commuter infrastructure make it effective for polished cultural launches and visible media support.",
    },
    "Brixton": {
        "brandFit": "medium",
        "description": "Grassroots nightlife and youth audience zone. Good for live programming, street presence, and community-led moments that need energy over polish.",
    },
    "Peckham": {
        "brandFit": "medium",
        "description": "Emerging creative cluster with a strong independent edge. Better for tastemaker activation and creator credibility than pure retail conversion.",
    },
    "Stratford": {
        "brandFit": "medium",
        "description": "Retail-and-OOH leverage zone for East London. Westfield-scale footfall and premium screen inventory make it the strongest awareness-to-conversion corridor outside the West End.",
    },
}

RETAIL_ANCHORS = [
    {"name": "Selfridges Oxford Street", "lat": 51.5146, "lng": -0.1528, "type": "department_store"},
    {"name": "John Lewis Oxford Street", "lat": 51.5145, "lng": -0.1463, "type": "department_store"},
    {"name": "Carnaby Street", "lat": 51.5138, "lng": -0.1396, "type": "retail_corridor"},
    {"name": "Covent Garden Piazza", "lat": 51.5118, "lng": -0.1242, "type": "retail_corridor"},
    {"name": "Coal Drops Yard", "lat": 51.5359, "lng": -0.1262, "type": "shopping_destination"},
    {"name": "Westfield London", "lat": 51.5072, "lng": -0.2202, "type": "shopping_destination"},
    {"name": "Westfield Stratford City", "lat": 51.5434, "lng": -0.0064, "type": "shopping_destination"},
    {"name": "Harrods", "lat": 51.4995, "lng": -0.1631, "type": "department_store"},
]

DESIGN_SCHOOLS = [
    {"name": "Central Saint Martins", "lat": 51.5351, "lng": -0.1257},
    {"name": "Chelsea College of Arts", "lat": 51.4948, "lng": -0.1072},
    {"name": "London College of Fashion, East Bank", "lat": 51.5429, "lng": -0.0100},
    {"name": "Goldsmiths, University of London", "lat": 51.4744, "lng": -0.0367},
    {"name": "Royal College of Art, Kensington", "lat": 51.5008, "lng": -0.1775},
    {"name": "Ravensbourne University London", "lat": 51.5017, "lng": 0.0035},
]

AGENCY_ANCHORS = [
    ("West End", 51.5138, -0.1362),
    ("Shoreditch", 51.5244, -0.0781),
    ("King's Cross", 51.5294, -0.1229),
    ("South Bank", 51.5054, -0.1046),
]

SEED_OOH_COORDINATES = {
    "Meridian Steps @ Westfield Stratford City": (51.5432, -0.0065),
    "Northern Ticket Hall @ Westfield Stratford City": (51.5440, -0.0058),
    "Skyline @ Westfield Stratford City": (51.5437, -0.0077),
    "The Street Totem @ Westfield Stratford City": (51.5427, -0.0071),
    "P10, Hammersmith Broadway": (51.4921, -0.2237),
    "The Arrival @ Heathrow Terminal 5": (51.4700, -0.4902),
    "The Eastern Lights, A13": (51.5165, 0.0425),
    "The Gateway @ Westfield London": (51.5074, -0.2208),
    "The Southern Terrace @ Westfield London": (51.5070, -0.2191),
    "The Screen @ Finchley Road, London": (51.5461, -0.1801),
    "The Screen @ High Street Kensington": (51.5008, -0.1935),
    "The Screen on Carnaby": (51.5137, -0.1394),
    "The Screen on Leicester Square": (51.5104, -0.1304),
    "The Screens @ Canary Wharf, Crossrail Place (Landscape)": (51.5053, -0.0195),
    "The Screens @ Canary Wharf, Crossrail Place (Portraits)": (51.5050, -0.0187),
    "The Screens @ Canary Wharf, Reuters Plaza (Landscape)": (51.5041, -0.0199),
    "The Screens @ Canary Wharf, Reuters Plaza (Portraits)": (51.5036, -0.0192),
    "Piccadilly Lights": (51.5101, -0.1338),
    "Tower Bridge DM6": (51.5007, -0.0759),
    "Bermondsey D96 Northbound": (51.4898, -0.0741),
    "Richmond Portrait": (51.4618, -0.2871),
    "Hendon DM6": (51.5748, -0.2283),
    "The One Knightsbridge": (51.4993, -0.1641),
    "The Landmark @ London Bridge": (51.5046, -0.0860),
}

TRANSIT_DEFAULTS = {
    "ubahn_poster": 20_000,
    "ubahn_special": 28_000,
    "bridge_banner": 55_000,
    "street_furniture": 15_000,
}

TRANSIT_WEIGHTS = {
    "piccadilly_lights": 180_000,
    "the_screen_on_leicester_square": 145_000,
    "the_screen_on_carnaby": 105_000,
    "meridian_steps_westfield_stratford_city": 110_000,
    "northern_ticket_hall_westfield_stratford_city": 108_000,
    "skyline_westfield_stratford_city": 102_000,
    "the_street_totem_westfield_stratford_city": 72_000,
    "the_gateway_westfield_london": 98_000,
    "the_southern_terrace_westfield_london": 84_000,
    "the_screens_canary_wharf_crossrail_place_landscape": 96_000,
    "the_screens_canary_wharf_crossrail_place_portraits": 88_000,
    "the_screens_canary_wharf_reuters_plaza_landscape": 90_000,
    "the_screens_canary_wharf_reuters_plaza_portraits": 82_000,
    "the_landmark_london_bridge": 85_000,
    "tower_bridge_dm6": 74_000,
    "p10_hammersmith_broadway": 68_000,
    "the_screen_finchley_road_london": 54_000,
    "the_screen_high_street_kensington": 62_000,
    "the_arrival_heathrow_terminal_5": 115_000,
    "the_eastern_lights_a13": 72_000,
    "the_one_knightsbridge": 78_000,
    "bermondsey_d96_northbound": 70_000,
    "richmond_portrait": 48_000,
    "hendon_dm6": 44_000,
}

DATA_SOURCE_META = {
    "retail": {
        "source": "Manual London retail anchor shortlist",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "medium",
        "note": "High-footfall retail destinations used as conversion anchors rather than an exhaustive partner inventory.",
    },
    "galleries": {
        "source": "GLA Cultural Infrastructure Map 2023 / gallery and museum resources",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "high",
        "note": "Open-data cultural inventory with geocoded venue records across London.",
    },
    "agencies": {
        "source": "London agency starter sheet plus inferred studio placement",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "low",
        "note": "Agency names come from a starter shortlist and use deterministic central-London placement where addresses were not provided.",
    },
    "coworking": {
        "source": "GLA Cultural Infrastructure Map 2023 / workspaces and third places",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "high",
        "note": "Creative coworking, workspaces, and public third places combined as creator-infrastructure signal.",
    },
    "venues": {
        "source": "GLA Cultural Infrastructure Map 2023 / theatres and nightlife resources",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "high",
        "note": "Venue layer combines theatre, nightlife, and museum-adjacent cultural destinations.",
    },
    "schools": {
        "source": "Manual London design-school shortlist",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "medium",
        "note": "Key design and arts campuses used as youth and creator touchpoints.",
    },
    "ubahn_poster": {
        "source": "London OOH proxy workbook / poster and transport panel buckets",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "medium",
        "note": "Deduped planning-proxy inventory mapped to poster-like transit and shelter formats.",
    },
    "ubahn_special": {
        "source": "London OOH proxy workbook / premium digital screens",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "medium",
        "note": "Digital screen and premium operator placements with heuristic weighting on flagship sites.",
    },
    "bridge_banner": {
        "source": "London OOH proxy workbook / large-format static inventory",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "medium",
        "note": "Billboards, hoardings, and banner wraps used as the large-format corridor layer.",
    },
    "street_furniture": {
        "source": "London OOH proxy workbook / kiosk, totem, and street-furniture formats",
        "updatedAt": DATA_REFRESH_DATE,
        "confidence": "medium",
        "note": "Street-level OOH proxy focused on kiosks, phone hubs, totems, and signage surfaces.",
    },
}

MARKET_META = {
    "code": "LDN",
    "city": "London",
    "locationLabel": "LONDON, UK",
    "reportTitle": "London location intelligence report",
    "mapTitle": "London Signal Map",
    "center": (51.5074, -0.1278),
    "coordsLabel": "LAT: 51.5074° N | LNG: 0.1278° W",
    "retailLabel": "Retail Anchors",
    "agenciesLabel": "Agencies + Studios",
    "venuesLabel": "Venues + Nightlife",
    "oohLabels": {
        "ubahn_poster": "Transit Posters",
        "ubahn_special": "Digital Screens",
        "bridge_banner": "Large Format",
        "street_furniture": "Street Furniture",
    },
    "oohModeBlurb": "Maps 2,500+ London OOH surfaces: transit posters, digital screens, large-format sites, and street furniture.",
    "hashtag": "#NothingLondon",
    "searchRegion": "UK",
}


def point_in_polygon(lat: float, lng: float, polygon: list[tuple[float, float]]) -> bool:
    inside = False
    j = len(polygon) - 1
    for i, (lat_i, lng_i) in enumerate(polygon):
        lat_j, lng_j = polygon[j]
        intersects = ((lng_i > lng) != (lng_j > lng)) and (
            lat < ((lat_j - lat_i) * (lng - lng_i)) / ((lng_j - lng_i) or 1e-12) + lat_i
        )
        if intersects:
            inside = not inside
        j = i
    return inside


def polygon_bounds(polygon: list[tuple[float, float]]) -> list[list[float]]:
    lats = [lat for lat, _ in polygon]
    lngs = [lng for _, lng in polygon]
    return [[round(min(lats), 4), round(min(lngs), 4)], [round(max(lats), 4), round(max(lngs), 4)]]


ZONE_CENTROIDS = {
    name: (
        sum(lat for lat, _ in polygon) / len(polygon),
        sum(lng for _, lng in polygon) / len(polygon),
    )
    for name, polygon in ZONE_POLYGONS.items()
}


def assign_zone(lat: float, lng: float) -> str:
    for zone_name in ZONE_ORDER:
        if point_in_polygon(lat, lng, ZONE_POLYGONS[zone_name]):
            return zone_name

    nearest_zone = ZONE_ORDER[0]
    nearest_distance = float("inf")
    for zone_name, (zone_lat, zone_lng) in ZONE_CENTROIDS.items():
        distance = (lat - zone_lat) ** 2 + (lng - zone_lng) ** 2
        if distance < nearest_distance:
            nearest_distance = distance
            nearest_zone = zone_name
    return nearest_zone


def stable_jitter(seed: str, lat_scale: float = 0.0032, lng_scale: float = 0.0046) -> tuple[float, float]:
    digest = hashlib.sha256(seed.encode("utf-8")).digest()
    lat_offset = ((digest[0] / 255) - 0.5) * 2 * lat_scale
    lng_offset = ((digest[1] / 255) - 0.5) * 2 * lng_scale
    return lat_offset, lng_offset


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def category_bucket(category_group: str, category: str) -> str | None:
    if category_group in {"Galleries", "Galleries & Museums"}:
        return "galleries"
    if category_group in {"Workspaces", "Third places"}:
        return "coworking"
    if category_group in {"Venues", "Nightlife"}:
        return "venues"
    if category_group == "Industry spaces":
        return "agencies"
    if category == "Fashion & design":
        return "agencies"
    return None


def ooh_bucket(media_type: str) -> str:
    if media_type in {"Digital / panel / screen", "Digital billboard / screen", "Digital OOH Screen", "Digital Roadside Billboard"}:
        return "ubahn_special"
    if media_type in {"Poster panel", "Advertising panel", "Bus shelter ad panel"}:
        return "ubahn_poster"
    if media_type in {"Billboard / hoarding", "Banner / wrap"}:
        return "bridge_banner"
    return "street_furniture"


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def extract_transit_node_key(name: str) -> str:
    ascii_name = clean_text(name)
    return normalize_key(ascii_name)


def sort_points(points: list[dict]) -> list[dict]:
    return sorted(points, key=lambda item: (item.get("area", ""), item["name"]))


def dedupe_points(points: list[dict], category: str) -> list[dict]:
    seen = set()
    unique: list[dict] = []
    for point in points:
        key = (
            category,
            point["name"].lower(),
            round(point["lat"], 6),
            round(point["lng"], 6),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(point)
    return unique


def build_manual_point(entry: dict, extra: dict | None = None) -> dict:
    zone = assign_zone(entry["lat"], entry["lng"])
    point = {
        "name": entry["name"],
        "lat": round(entry["lat"], 6),
        "lng": round(entry["lng"], 6),
        "area": zone,
    }
    if "type" in entry:
        point["type"] = entry["type"]
    if extra:
        point.update(extra)
    return point


def parse_creative_workbook() -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(CREATIVE_WORKBOOK, read_only=True)
    ws = wb["Master_Creative_Places"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_map = {name: index for index, name in enumerate(header)}

    buckets: dict[str, list[dict]] = defaultdict(list)

    for row in rows:
        if not row or all(cell is None for cell in row):
            continue

        name = clean_text(row[header_map["Name"]])
        category_group = clean_text(row[header_map["Category_Group"]])
        category = clean_text(row[header_map["Category"]])
        lat = row[header_map["Latitude"]]
        lng = row[header_map["Longitude"]]
        if not name or lat is None or lng is None:
            continue

        bucket = category_bucket(category_group, category)
        if not bucket:
            continue

        zone = assign_zone(float(lat), float(lng))
        point = {
            "name": name,
            "lat": round(float(lat), 6),
            "lng": round(float(lng), 6),
            "area": zone,
            "district": clean_text(row[header_map["Borough"]]) or None,
            "plz": clean_text(row[header_map["Postcode"]]) or None,
        }
        buckets[bucket].append({key: value for key, value in point.items() if value is not None})

    agency_sheet = wb["Marketing_Agencies_Starter"]
    agency_rows = agency_sheet.iter_rows(values_only=True)
    agency_header = next(agency_rows)
    agency_map = {name: index for index, name in enumerate(agency_header)}

    for index, row in enumerate(agency_rows):
        name = clean_text(row[agency_map["Agency_Name"]])
        if not name:
            continue
        zone_name, base_lat, base_lng = AGENCY_ANCHORS[index % len(AGENCY_ANCHORS)]
        lat_offset, lng_offset = stable_jitter(name, 0.0045, 0.0065)
        point = {
            "name": name,
            "lat": round(base_lat + lat_offset, 6),
            "lng": round(base_lng + lng_offset, 6),
            "area": zone_name,
            "district": "Inferred central placement",
        }
        buckets["agencies"].append(point)

    wb.close()

    buckets["retail"] = [build_manual_point(anchor) for anchor in RETAIL_ANCHORS]
    buckets["schools"] = [build_manual_point(school) for school in DESIGN_SCHOOLS]

    return {
        category: sort_points(dedupe_points(points, category))
        for category, points in {
            "retail": buckets["retail"],
            "galleries": buckets["galleries"],
            "agencies": buckets["agencies"],
            "coworking": buckets["coworking"],
            "venues": buckets["venues"],
            "schools": buckets["schools"],
        }.items()
    }


def looks_like_planning_ref(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}/[A-Z0-9]+/[A-Z]", value))


def build_ooh_name(location_name: str, address: str, media_type: str) -> str:
    clean_location = clean_text(location_name)
    clean_address = clean_text(address)
    if clean_location and not looks_like_planning_ref(clean_location):
        return clean_location
    if clean_address:
        return clean_address
    return clean_location or media_type


def parse_ooh_workbook() -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(OOH_WORKBOOK, read_only=True)
    ws = wb["OOH_Likely_Unique_Sites"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_map = {name: index for index, name in enumerate(header)}

    buckets: dict[str, list[dict]] = defaultdict(list)

    for row in rows:
        if not row or all(cell is None for cell in row):
            continue

        lat = row[header_map["Latitude"]]
        lng = row[header_map["Longitude"]]
        media_type = clean_text(row[header_map["Media_Type"]])
        if lat is None or lng is None or not media_type:
            continue

        bucket = ooh_bucket(media_type)
        zone = assign_zone(float(lat), float(lng))
        name = build_ooh_name(
            clean_text(row[header_map["Location_Name"]]),
            clean_text(row[header_map["Area_or_Address"]]),
            media_type,
        )
        point = {
            "name": name,
            "lat": round(float(lat), 6),
            "lng": round(float(lng), 6),
            "area": zone,
            "plz": clean_text(row[header_map["Postcode"]]) or None,
            "media_type": media_type,
            "source_type": clean_text(row[header_map["Source_Type"]]) or None,
        }
        buckets[bucket].append({key: value for key, value in point.items() if value is not None})

    seed_ws = wb["Seed_Operator_Spots"]
    seed_rows = seed_ws.iter_rows(values_only=True)
    seed_header = next(seed_rows)
    seed_map = {name: index for index, name in enumerate(seed_header)}

    for row in seed_rows:
        name = clean_text(row[seed_map["Location_Name"]])
        media_type = clean_text(row[seed_map["Media_Type"]])
        coords = SEED_OOH_COORDINATES.get(name)
        if not name or not coords:
            continue

        bucket = ooh_bucket(media_type)
        lat, lng = coords
        zone = assign_zone(lat, lng)
        point = {
            "name": name,
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "area": zone,
            "media_type": media_type,
            "source_type": clean_text(row[seed_map["Operator"]]) or None,
        }
        buckets[bucket].append(point)

    wb.close()

    return {
        category: sort_points(dedupe_points(points, category))
        for category, points in {
            "ubahn_poster": buckets["ubahn_poster"],
            "ubahn_special": buckets["ubahn_special"],
            "bridge_banner": buckets["bridge_banner"],
            "street_furniture": buckets["street_furniture"],
        }.items()
    }


def get_impressions(point: dict, category: str) -> int:
    lookup_key = extract_transit_node_key(point["name"])
    return TRANSIT_WEIGHTS.get(lookup_key, TRANSIT_DEFAULTS[category])


def build_neighbourhoods(locations: dict[str, list[dict]], ooh_locations: dict[str, list[dict]]) -> dict:
    stats_by_zone: dict[str, dict[str, int]] = {
        zone_name: {
            "galleries": 0,
            "agencies": 0,
            "venues": 0,
            "coworking": 0,
            "retail": 0,
            "schools": 0,
            "ubahn_poster": 0,
            "ubahn_special": 0,
            "bridge_banner": 0,
            "street_furniture": 0,
        }
        for zone_name in ZONE_ORDER
    }
    impressions_by_zone = {
        zone_name: {
            "ubahn_poster": 0,
            "ubahn_special": 0,
            "bridge_banner": 0,
            "street_furniture": 0,
            "total": 0,
        }
        for zone_name in ZONE_ORDER
    }

    for category, points in locations.items():
        for point in points:
            stats_by_zone[point["area"]][category] += 1

    for category, points in ooh_locations.items():
        for point in points:
            zone_name = point["area"]
            stats_by_zone[zone_name][category] += 1
            impression_value = get_impressions(point, category)
            impressions_by_zone[zone_name][category] += impression_value
            impressions_by_zone[zone_name]["total"] += impression_value

    raw_scores: dict[str, dict[str, float]] = {}
    for zone_name in ZONE_ORDER:
        stats = stats_by_zone[zone_name]
        impressions = impressions_by_zone[zone_name]
        raw_scores[zone_name] = {
            "cultural": (
                stats["galleries"] * 1.35
                + stats["agencies"] * 0.95
                + stats["venues"] * 1.2
                + stats["coworking"] * 0.9
                + stats["schools"] * 0.75
            ),
            "retail": (
                stats["retail"] * 1.55
                + stats["galleries"] * 0.55
                + stats["venues"] * 0.5
                + impressions["ubahn_special"] / 240_000
            ),
            "creator": (
                stats["agencies"] * 1.35
                + stats["coworking"] * 1.25
                + stats["galleries"] * 0.55
                + stats["schools"] * 0.7
            ),
            "guerrilla": (
                stats["venues"] * 1.4
                + stats["schools"] * 1.15
                + stats["street_furniture"] * 0.1
                + stats["bridge_banner"] * 0.18
            ),
            "ooh": (
                impressions["ubahn_poster"] / 80_000
                + impressions["ubahn_special"] / 70_000
                + impressions["bridge_banner"] / 85_000
                + impressions["street_furniture"] / 95_000
            ),
        }

    score_ranges = {
        "cultural": (42, 90),
        "retail": (40, 88),
        "creator": (40, 91),
        "guerrilla": (38, 90),
        "ooh": (36, 92),
    }

    scaled_scores: dict[str, dict[str, int]] = {zone_name: {} for zone_name in ZONE_ORDER}
    for mode, (score_floor, score_ceiling) in score_ranges.items():
        values = [raw_scores[zone_name][mode] for zone_name in ZONE_ORDER]
        min_value = min(values)
        max_value = max(values)
        for zone_name in ZONE_ORDER:
            current = raw_scores[zone_name][mode]
            if math.isclose(max_value, min_value):
                scaled = round((score_floor + score_ceiling) / 2)
            else:
                scaled = round(
                    score_floor + ((current - min_value) / (max_value - min_value)) * (score_ceiling - score_floor)
                )
            scaled_scores[zone_name][mode] = int(scaled)

    neighbourhoods = []
    for zone_name in ZONE_ORDER:
        neighbourhoods.append(
            {
                "name": zone_name,
                "scores": scaled_scores[zone_name],
                "stats": stats_by_zone[zone_name],
                "description": ZONE_COPY[zone_name]["description"],
                "brandFit": ZONE_COPY[zone_name]["brandFit"],
                "bounds": polygon_bounds(ZONE_POLYGONS[zone_name]),
            }
        )

    return {"neighbourhoods": neighbourhoods}


def ts_value(value, indent: int = 0) -> str:
    pad = " " * indent
    if isinstance(value, str):
        return json.dumps(value)
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return repr(value)
    if isinstance(value, tuple):
        inner = ", ".join(ts_value(item, indent) for item in value)
        return f"[{inner}]"
    if isinstance(value, list):
        if not value:
            return "[]"
        next_indent = indent + 2
        inner = ",\n".join(f'{" " * next_indent}{ts_value(item, next_indent)}' for item in value)
        return "[\n" + inner + f"\n{pad}]"
    if isinstance(value, dict):
        if not value:
            return "{}"
        next_indent = indent + 2
        lines = []
        for key, item in value.items():
            safe_key = key if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", key) else json.dumps(key)
            lines.append(f'{" " * next_indent}{safe_key}: {ts_value(item, next_indent)}')
        return "{\n" + ",\n".join(lines) + f"\n{pad}" + "}"
    raise TypeError(f"Unsupported type for TS conversion: {type(value)!r}")


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def write_zone_metadata() -> None:
    content = (
        'import { ConfidenceLevel, LocationCategory } from "@/components/types";\n\n'
        f'export const DATA_REFRESH_DATE = "{DATA_REFRESH_DATE}";\n\n'
        "export const ZONE_POLYGONS: Record<string, [number, number][]> = "
        + ts_value({zone: [[lat, lng] for lat, lng in polygon] for zone, polygon in ZONE_POLYGONS.items()}, 0)
        + ";\n\n"
        "export const DATA_SOURCE_META: Record<\n"
        "  LocationCategory,\n"
        "  {\n"
        "    confidence: ConfidenceLevel;\n"
        "    note: string;\n"
        "    source: string;\n"
        "    updatedAt: string;\n"
        "  }\n"
        "> = "
        + ts_value(DATA_SOURCE_META, 0)
        + ";\n"
    )
    ZONE_METADATA_OUTPUT.write_text(content, encoding="utf-8")


def write_transit_weights() -> None:
    content = (
        'import { OOHCategory } from "@/components/types";\n\n'
        "export const TRANSIT_DEFAULTS: Record<OOHCategory, number> = "
        + ts_value(TRANSIT_DEFAULTS, 0)
        + ";\n\n"
        "// Heuristic daily impression estimates for premium London OOH sites.\n"
        "export const TRANSIT_WEIGHTS: Record<string, number> = "
        + ts_value(TRANSIT_WEIGHTS, 0)
        + ";\n"
    )
    TRANSIT_WEIGHTS_OUTPUT.write_text(content, encoding="utf-8")


def write_market_meta() -> None:
    content = (
        "export const MARKET = {\n"
        f'  code: {json.dumps(MARKET_META["code"])},\n'
        f'  city: {json.dumps(MARKET_META["city"])},\n'
        f'  locationLabel: {json.dumps(MARKET_META["locationLabel"])},\n'
        f'  reportTitle: {json.dumps(MARKET_META["reportTitle"])},\n'
        f'  mapTitle: {json.dumps(MARKET_META["mapTitle"])},\n'
        f"  center: [{MARKET_META['center'][0]}, {MARKET_META['center'][1]}] as [number, number],\n"
        f'  coordsLabel: {json.dumps(MARKET_META["coordsLabel"])},\n'
        f'  retailLabel: {json.dumps(MARKET_META["retailLabel"])},\n'
        f'  agenciesLabel: {json.dumps(MARKET_META["agenciesLabel"])},\n'
        f'  venuesLabel: {json.dumps(MARKET_META["venuesLabel"])},\n'
        f"  oohLabels: {ts_value(MARKET_META['oohLabels'], 2)},\n"
        f'  oohModeBlurb: {json.dumps(MARKET_META["oohModeBlurb"])},\n'
        f'  hashtag: {json.dumps(MARKET_META["hashtag"])},\n'
        f'  searchRegion: {json.dumps(MARKET_META["searchRegion"])},\n'
        "} as const;\n"
    )
    MARKET_OUTPUT.write_text(content, encoding="utf-8")


def print_summary(locations: dict[str, list[dict]], ooh_locations: dict[str, list[dict]], neighbourhoods: dict) -> None:
    print("London market data generated")
    print("-" * 60)
    for category, points in locations.items():
        print(f"{category:>16}: {len(points)}")
    for category, points in ooh_locations.items():
        print(f"{category:>16}: {len(points)}")
    print("-" * 60)
    for zone in neighbourhoods["neighbourhoods"]:
        score_line = ", ".join(f"{mode} {score}" for mode, score in zone["scores"].items())
        print(f'{zone["name"]:>16}: {score_line}')


def main() -> None:
    locations = parse_creative_workbook()
    ooh_locations = parse_ooh_workbook()
    neighbourhoods = build_neighbourhoods(locations, ooh_locations)

    write_json(LOCATIONS_OUTPUT, locations)
    write_json(OOH_OUTPUT, ooh_locations)
    write_json(NEIGHBOURHOODS_OUTPUT, neighbourhoods)
    write_zone_metadata()
    write_transit_weights()
    write_market_meta()
    print_summary(locations, ooh_locations, neighbourhoods)


if __name__ == "__main__":
    main()
