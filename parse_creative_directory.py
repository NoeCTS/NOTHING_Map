#!/usr/bin/env python3
"""
Parse Berlin Creative Ecosystem Directory Excel file and generate JSON output
for the Ground Signal app.
"""

import json
import math
import random
import time
from pathlib import Path

import openpyxl
import requests

# Configuration
EXCEL_PATH = "/Users/noeelamine/Documents/Playground/Berlin_Creative_Ecosystem_Directory_with_More_Agencies.xlsx"
OUTPUT_PATH = "/Users/noeelamine/Documents/Playground/data/creative_spaces.json"

# Category mapping
CATEGORY_MAP = {
    "Galleries & Museums": "galleries",
    "Marketing Agencies": "agencies",
    "Theatres & Performance": "venues",
    "Funded Cultural Venues & Programs": "venues",
    "Libraries & Archives": "coworking",
    "Memorials, Palaces & Gardens": "venues",
}

# Comprehensive PLZ centroids for Berlin
PLZ_CENTROIDS = {
    "10115": (52.5316, 13.3881),
    "10117": (52.5163, 13.3883),
    "10119": (52.5303, 13.4089),
    "10178": (52.5219, 13.4097),
    "10179": (52.5134, 13.4198),
    "10243": (52.5103, 13.4545),
    "10245": (52.5006, 13.4543),
    "10247": (52.5156, 13.4638),
    "10249": (52.5253, 13.4456),
    "10317": (52.5089, 13.4861),
    "10318": (52.4856, 13.4921),
    "10319": (52.4958, 13.5069),
    "10365": (52.5208, 13.4939),
    "10367": (52.5233, 13.4753),
    "10405": (52.5347, 13.4269),
    "10407": (52.5392, 13.4445),
    "10409": (52.5481, 13.4356),
    "10435": (52.5389, 13.4128),
    "10437": (52.5467, 13.4150),
    "10439": (52.5528, 13.4033),
    "10551": (52.5294, 13.3419),
    "10553": (52.5236, 13.3394),
    "10555": (52.5197, 13.3264),
    "10557": (52.5203, 13.3609),
    "10559": (52.5311, 13.3577),
    "10585": (52.5167, 13.3014),
    "10587": (52.5156, 13.3233),
    "10589": (52.5281, 13.3031),
    "10623": (52.5078, 13.3292),
    "10625": (52.5083, 13.3181),
    "10627": (52.5064, 13.3061),
    "10629": (52.5008, 13.3136),
    "10707": (52.4928, 13.3136),
    "10709": (52.4894, 13.3033),
    "10711": (52.4944, 13.2892),
    "10713": (52.4806, 13.3117),
    "10715": (52.4803, 13.3331),
    "10717": (52.4869, 13.3275),
    "10719": (52.4981, 13.3256),
    "10777": (52.4992, 13.3472),
    "10779": (52.4939, 13.3511),
    "10781": (52.4933, 13.3647),
    "10783": (52.4997, 13.3711),
    "10785": (52.5053, 13.3658),
    "10787": (52.5056, 13.3411),
    "10789": (52.5042, 13.3306),
    "10823": (52.4853, 13.3578),
    "10825": (52.4817, 13.3458),
    "10827": (52.4769, 13.3544),
    "10829": (52.4756, 13.3672),
    "10961": (52.4897, 13.3919),
    "10963": (52.5011, 13.3847),
    "10965": (52.4833, 13.3931),
    "10967": (52.4858, 13.4092),
    "10969": (52.5019, 13.4078),
    "10997": (52.5017, 13.4333),
    "10999": (52.4967, 13.4231),
    "12043": (52.4806, 13.4311),
    "12045": (52.4803, 13.4483),
    "12047": (52.4858, 13.4328),
    "12049": (52.4764, 13.4222),
    "12051": (52.4692, 13.4339),
    "12053": (52.4742, 13.4450),
    "12055": (52.4689, 13.4508),
    "12057": (52.4614, 13.4414),
    "12059": (52.4672, 13.4617),
    "12099": (52.4622, 13.4231),
    "12157": (52.4658, 13.3319),
    "12159": (52.4728, 13.3394),
    "12161": (52.4692, 13.3217),
    "12163": (52.4631, 13.3089),
    "12165": (52.4567, 13.3069),
    "12167": (52.4539, 13.3206),
    "12169": (52.4533, 13.3356),
    "13086": (52.5586, 13.4219),
    "13088": (52.5667, 13.4453),
    "13089": (52.5744, 13.4464),
    # Additional PLZ from the data
    "13353": (52.5450, 13.3550),  # Wedding
    "13403": (52.5800, 13.3170),  # Reinickendorf
    "12439": (52.4530, 13.5180),  # Schoneweide
    "14195": (52.4570, 13.2850),  # Dahlem
    "14467": (52.3950, 13.0617),  # Potsdam area
    "10969": (52.5019, 13.4078),  # Kreuzberg
    "12205": (52.4333, 13.3150),  # Lichterfelde
}

# Neighbourhood zone definitions
ZONES = {
    "Kreuzberg": {"lat_min": 52.48, "lat_max": 52.51, "lng_min": 13.38, "lng_max": 13.45},
    "Mitte": {"lat_min": 52.515, "lat_max": 52.535, "lng_min": 13.37, "lng_max": 13.42},
    "Friedrichshain": {"lat_min": 52.505, "lat_max": 52.52, "lng_min": 13.43, "lng_max": 13.48},
    "Neukolln": {"lat_min": 52.47, "lat_max": 52.49, "lng_min": 13.42, "lng_max": 13.47},
    "Charlottenburg": {"lat_min": 52.5, "lat_max": 52.52, "lng_min": 13.29, "lng_max": 13.34},
    "Prenzlauer Berg": {"lat_min": 52.535, "lat_max": 52.555, "lng_min": 13.4, "lng_max": 13.44},
}

# Zone centroids for fallback
ZONE_CENTROIDS = {
    name: (
        (zone["lat_min"] + zone["lat_max"]) / 2,
        (zone["lng_min"] + zone["lng_max"]) / 2
    )
    for name, zone in ZONES.items()
}


def get_plz_fallback(postal_code: str) -> tuple[float, float] | None:
    """Get PLZ centroid with random jitter."""
    plz = str(postal_code).strip()[:5] if postal_code else None
    if plz and plz in PLZ_CENTROIDS:
        lat, lng = PLZ_CENTROIDS[plz]
        # Add small random jitter (roughly 50-100m)
        jitter_lat = random.uniform(-0.001, 0.001)
        jitter_lng = random.uniform(-0.001, 0.001)
        return lat + jitter_lat, lng + jitter_lng
    return None


def assign_zone(lat: float, lng: float) -> str:
    """Assign a neighbourhood zone based on coordinates."""
    # First, check if point is within any zone
    for zone_name, bounds in ZONES.items():
        if (bounds["lat_min"] <= lat <= bounds["lat_max"] and
            bounds["lng_min"] <= lng <= bounds["lng_max"]):
            return zone_name
    
    # If outside all zones, find nearest centroid
    min_dist = float("inf")
    nearest_zone = "Mitte"  # Default fallback
    
    for zone_name, (clat, clng) in ZONE_CENTROIDS.items():
        dist = math.sqrt((lat - clat) ** 2 + (lng - clng) ** 2)
        if dist < min_dist:
            min_dist = dist
            nearest_zone = zone_name
    
    return nearest_zone


def get_cell_value(row, header_map, *possible_names, default=None):
    """Get cell value by trying multiple possible column names."""
    for name in possible_names:
        if name in header_map:
            idx = header_map[name]
            if idx < len(row) and row[idx] is not None:
                return row[idx]
    return default


def parse_excel() -> list:
    """Parse the Master_Creative_Places sheet from the Excel file."""
    print(f"Loading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    
    all_entries = []
    
    # Process Master_Creative_Places sheet only (it contains all data)
    if "Master_Creative_Places" not in wb.sheetnames:
        print("ERROR: Master_Creative_Places sheet not found!")
        wb.close()
        return []
    
    sheet = wb["Master_Creative_Places"]
    rows = list(sheet.iter_rows(values_only=True))
    
    if not rows:
        wb.close()
        return []
    
    header = rows[0]
    header_map = {str(h).strip(): i for i, h in enumerate(header) if h}
    
    print(f"\nProcessing sheet: Master_Creative_Places")
    print(f"  Headers: {list(header_map.keys())}")
    print(f"  Rows (excluding header): {len(rows) - 1}")
    
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        
        name = get_cell_value(row, header_map, "Name")
        if not name:
            continue
        
        category = get_cell_value(row, header_map, "Category", default="Unknown")
        street = get_cell_value(row, header_map, "Street_Address", "Address", default="")
        postal = get_cell_value(row, header_map, "Postal_Code", default="")
        lat = get_cell_value(row, header_map, "Latitude")
        lng = get_cell_value(row, header_map, "Longitude")
        
        # Convert coordinates
        try:
            lat = float(lat) if lat is not None else None
            lng = float(lng) if lng is not None else None
        except (ValueError, TypeError):
            lat = None
            lng = None
        
        all_entries.append({
            "name": str(name).strip(),
            "category": str(category).strip() if category else "Unknown",
            "street": str(street).strip() if street else "",
            "postal": str(postal).strip() if postal else "",
            "lat": lat,
            "lng": lng,
        })
    
    wb.close()
    
    print(f"\nTotal entries parsed: {len(all_entries)}")
    return all_entries


def geocode_entries(entries: list) -> tuple[list, dict]:
    """Geocode entries missing coordinates using PLZ fallback."""
    stats = {
        "had_coords": 0,
        "plz_fallback": 0,
        "failed": 0,
    }
    
    needs_geocoding = [e for e in entries if e["lat"] is None or e["lng"] is None]
    print(f"  Entries needing geocoding: {len(needs_geocoding)}")
    
    for entry in entries:
        if entry["lat"] is not None and entry["lng"] is not None:
            stats["had_coords"] += 1
            continue
        
        # Use PLZ fallback (Nominatim is blocked)
        coords = get_plz_fallback(entry["postal"])
        if coords:
            entry["lat"], entry["lng"] = coords
            stats["plz_fallback"] += 1
            continue
        
        # If no PLZ, use Berlin center with jitter for agencies
        # Distribute them across central Berlin neighborhoods
        if entry["category"] == "Marketing Agencies":
            # Pick a random central Berlin location with spread
            base_locations = [
                (52.5200, 13.4050),  # Mitte
                (52.5000, 13.4200),  # Kreuzberg
                (52.5100, 13.3300),  # Charlottenburg
                (52.5400, 13.4150),  # Prenzlauer Berg
                (52.5100, 13.4500),  # Friedrichshain
            ]
            base_lat, base_lng = random.choice(base_locations)
            entry["lat"] = base_lat + random.uniform(-0.01, 0.01)
            entry["lng"] = base_lng + random.uniform(-0.01, 0.01)
        else:
            # Generic Berlin center fallback
            entry["lat"] = 52.52 + random.uniform(-0.02, 0.02)
            entry["lng"] = 13.405 + random.uniform(-0.02, 0.02)
        
        stats["failed"] += 1
        print(f"  No PLZ for: {entry['name'][:50]} - using Berlin fallback")
    
    return entries, stats


def build_output(entries: list) -> dict:
    """Build the output JSON structure."""
    output = {
        "galleries": [],
        "agencies": [],
        "venues": [],
        "coworking": [],
    }
    
    zone_counts = {zone: 0 for zone in ZONES}
    category_counts = {cat: 0 for cat in output.keys()}
    unmapped_categories = set()
    
    for entry in entries:
        # Map category
        app_category = CATEGORY_MAP.get(entry["category"])
        if not app_category:
            unmapped_categories.add(entry["category"])
            app_category = "venues"
        
        # Assign zone
        zone = assign_zone(entry["lat"], entry["lng"])
        
        # Create output entry
        point = {
            "name": entry["name"],
            "lat": round(entry["lat"], 6),
            "lng": round(entry["lng"], 6),
            "area": zone,
        }
        
        output[app_category].append(point)
        category_counts[app_category] += 1
        zone_counts[zone] += 1
    
    if unmapped_categories:
        print(f"\nUnmapped categories (defaulted to 'venues'): {unmapped_categories}")
    
    return output, category_counts, zone_counts


def main():
    print("=" * 60)
    print("Berlin Creative Ecosystem Directory Parser")
    print("=" * 60)
    
    # Step 1: Parse Excel
    print("\n[STEP 1] Parsing Excel sheets...")
    entries = parse_excel()
    
    if not entries:
        print("ERROR: No entries parsed from Excel file!")
        return
    
    # Step 2: Show category mapping
    print("\n[STEP 2] Category mapping:")
    for original, mapped in CATEGORY_MAP.items():
        print(f"  '{original}' -> '{mapped}'")
    
    # Step 3: Geocode missing coordinates
    print("\n[STEP 3] Geocoding missing coordinates...")
    entries, geocode_stats = geocode_entries(entries)
    
    # Step 4 & 5: Build output and assign zones
    print("\n[STEP 4 & 5] Building output JSON and assigning zones...")
    output, category_counts, zone_counts = build_output(entries)
    
    # Ensure output directory exists
    output_path = Path(OUTPUT_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Write JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    print(f"\nOutput written to: {OUTPUT_PATH}")
    
    # Step 6: Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    
    print("\n[Category Totals]")
    total = 0
    for cat, count in category_counts.items():
        print(f"  {cat}: {count}")
        total += count
    print(f"  TOTAL: {total}")
    
    print("\n[Zone Totals]")
    for zone, count in sorted(zone_counts.items(), key=lambda x: -x[1]):
        print(f"  {zone}: {count}")
    
    print("\n[Geocoding Stats]")
    print(f"  Already had coordinates: {geocode_stats['had_coords']}")
    print(f"  PLZ fallback used: {geocode_stats['plz_fallback']}")
    print(f"  Failed (Berlin center): {geocode_stats['failed']}")
    
    total_entries = sum(geocode_stats.values())
    success = geocode_stats['had_coords'] + geocode_stats['plz_fallback']
    success_rate = (success / total_entries * 100) if total_entries > 0 else 0
    print(f"\n  Geocoding success rate: {success_rate:.1f}%")
    
    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    main()
